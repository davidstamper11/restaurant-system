from fastapi import FastAPI
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Form, Request, status, UploadFile, File
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from app.models.db_base import Review, Restaurant
from app.models.authorized_reviewer import AuthorizedReviewer
from app.schemas.review_schema import ReviewCreateSchema, ReviewResponseSchema
from app.schemas.authorized_reviewer_schema import AuthorizedReviewerResponseSchema
from app.services.db_service import get_db
from passlib.context import CryptContext
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
import os
jinja_templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "..", "templates"))

router = APIRouter(prefix="/admin", tags=["admin"])

# Middleware for admin authentication
class AdminAuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.url.path in ("/admin/login", "/admin/"):
            return await call_next(request)
        session_id = request.cookies.get("session_id")
        if not session_id:
            return RedirectResponse(url="/admin/login", status_code=303)
        from app.services.db_service import get_db
        db = next(get_db())
        try:
            admin = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == int(session_id)).first()
            if not admin or not admin.is_admin_yn:
                return RedirectResponse(url="/admin/login", status_code=303)
        except Exception:
            return RedirectResponse(url="/admin/login", status_code=303)
        finally:
            db.close()
        return await call_next(request)
        # Check for valid session cookie
        session_id = request.cookies.get("session_id")
        if not session_id:
            return RedirectResponse(url="/admin/login", status_code=303)
        # Verify admin user
        from app.services.db_service import get_db
        db = next(get_db())
        try:
            admin = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == int(session_id)).first()
            if not admin or not admin.is_admin_yn:
                return RedirectResponse(url="/admin/login", status_code=303)
        except Exception:
            return RedirectResponse(url="/admin/login", status_code=303)
        finally:
            db.close()
        return await call_next(request)

# Existing routes unchanged

@router.get("/login", response_class=HTMLResponse)
async def admin_login_form(request: Request):
    template = jinja_templates.get_template("admin_login.html")
    return HTMLResponse(content=template.render())

@router.post("/login")
async def admin_login(email: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    admin = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.email == email, AuthorizedReviewer.is_admin_yn == True).first()
    if not admin or not pwd_context.verify(password, admin.password):
        raise HTTPException(status_code=403, detail="Forbidden")
    response = RedirectResponse(url="/admin/dashboard", status_code=303)
    response.set_cookie(key="session_id", value=str(admin.id), httponly=True, secure=True, samesite="Strict")
    return response

@router.get("/logout", response_class=HTMLResponse)
async def admin_logout():
    response = RedirectResponse(url="/admin/login", status_code=303)
    response.delete_cookie(key="session_id")
    return response
@router.get("/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    template = jinja_templates.get_template("admin_dashboard.html")
    return HTMLResponse(content=template.render({"request": request}))

# List Reviewers
@router.get("/reviewers", response_class=HTMLResponse)
async def list_reviewers(request: Request, db: Session = Depends(get_db)):
    reviewers = db.query(AuthorizedReviewer).all()
    return HTMLResponse(content=jinja_templates.get_template("admin_reviewers.html").render({"request": request, "reviewers": reviewers}))

# New Reviewer Form
@router.get("/reviewers/new", response_class=HTMLResponse)
async def new_reviewer_form(request: Request):
    return HTMLResponse(
        content=jinja_templates.get_template("admin_reviewer_form.html").render(
            request=request,
            action="/admin/reviewers",
            method="POST",
            reviewer=None,
        )
    )

# CREATE – process form submission
@router.post("/reviewers")
async def create_reviewer(
    name: str = Form(...),
    email: str = Form(...),
    is_admin_yn: bool = Form(False),
    password: str = Form(None),
    db: Session = Depends(get_db),
):
    new_reviewer = AuthorizedReviewer(
        name=name,
        email=email,
        is_admin_yn=is_admin_yn,
        password=pwd_context.hash(password) if password else None,
    )
    db.add(new_reviewer)
    db.commit()
    db.refresh(new_reviewer)
    return RedirectResponse(url="/admin/reviewers", status_code=status.HTTP_303_SEE_OTHER)

# UPDATE – process edit form submission
@router.post("/reviewers/{reviewer_id}")
async def update_reviewer(
    reviewer_id: int,
    name: str = Form(...),
    email: str = Form(...),
    is_admin_yn: bool = Form(False),
    password: str = Form(None),
    db: Session = Depends(get_db),
):
    reviewer = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")
    reviewer.name = name
    reviewer.email = email
    reviewer.is_admin_yn = is_admin_yn
    # Update password only if a new one is provided; otherwise keep existing
    if password:
        reviewer.password = pwd_context.hash(password)
    db.commit()
    return RedirectResponse(url="/admin/reviewers", status_code=status.HTTP_303_SEE_OTHER)

# Edit Reviewer Form
@router.get("/reviewers/{reviewer_id}/edit", response_class=HTMLResponse)
async def edit_reviewer_form(reviewer_id: int, request: Request, db: Session = Depends(get_db)):
    reviewer = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")
    return HTMLResponse(
        content=jinja_templates.get_template("admin_reviewer_form.html").render(
            request=request,
            action=f"/admin/reviewers/{reviewer_id}",
            method="POST",
            reviewer=reviewer,
        )
    )

# UPDATE – process edit form submission
@router.post("/reviewers/{reviewer_id}")
async def update_reviewer(
    reviewer_id: int,
    name: str = Form(...),
    email: str = Form(...),
    is_admin_yn: bool = Form(False),
    password: str = Form(None),
    db: Session = Depends(get_db),
):
    reviewer = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")
    reviewer.name = name
    reviewer.email = email
    reviewer.is_admin_yn = is_admin_yn
    # Only update password if a new one was provided
    if password:
        reviewer.password = pwd_context.hash(password)
    db.commit()
    return RedirectResponse(url="/admin/reviewers", status_code=status.HTTP_303_SEE_OTHER)

# DELETE – remove reviewer
@router.post("/reviewers/{reviewer_id}/delete")
async def delete_reviewer(reviewer_id: int, db: Session = Depends(get_db)):
    reviewer = db.query(AuthorizedReviewer).filter(AuthorizedReviewer.id == reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")
    db.delete(reviewer)
    db.commit()
    return RedirectResponse(url="/admin/reviewers", status_code=status.HTTP_303_SEE_OTHER)

# ----------------------------------------------------------------------
# Manage Reviews – list
# ----------------------------------------------------------------------
@router.get("/reviews", response_class=HTMLResponse)
async def admin_list_reviews(request: Request, db: Session = Depends(get_db)):
    reviews = db.query(Review).order_by(Review.id.desc()).all()
    return HTMLResponse(content=jinja_templates.get_template("admin_reviews.html").render({"request": request, "reviews": reviews}))

# ----------------------------------------------------------------------
# Delete Review
# ----------------------------------------------------------------------
@router.get("/reviews/delete/{review_id}")
async def admin_delete_review(review_id: int, db: Session = Depends(get_db)):
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    # delete associated photos if they exist
    for photo in (review.photo1, review.photo2):
        if photo:
            db.delete(photo)
    db.delete(review)
    db.commit()
    return RedirectResponse(url="/admin/reviews", status_code=status.HTTP_303_SEE_OTHER)

# ----------------------------------------------------------------------
# Edit Review – form (reuse public review_form template)
# ----------------------------------------------------------------------
@router.get("/reviews/edit/{review_id}", response_class=HTMLResponse)
async def admin_edit_review_form(request: Request, review_id: int, db: Session = Depends(get_db)):
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    return HTMLResponse(content=jinja_templates.get_template("review_form.html").render({
        "request": request,
        "review": review,
        "restaurant_id": review.restaurant_id,
        "edit_mode": True,
    }))

# ----------------------------------------------------------------------
# Edit Review – submit update
# ----------------------------------------------------------------------
@router.post("/reviews/edit/{review_id}")
async def admin_update_review(
    review_id: int,
    restaurant_id: int = Form(...),
    reviewer_email: str = Form(...),
    score: float = Form(...),
    comments: str = Form(None),
    price_subscore: int = Form(0),
    salsa_subscore: int = Form(0),
    tortillas_subscore: int = Form(0),
    condiments_subscore: int = Form(0),
    ambiance_subscore: int = Form(0),
    flavor_subscore: int = Form(0),
    service_subscore: int = Form(0),
    homemade_tortillas_yn: bool = Form(False),
    other_notes: str = Form(None),
    photo_file_1: UploadFile = File(None),
    photo_file_2: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    from app.models.db_base import Photo
    import os, shutil, uuid
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    # update scalar fields
    review.restaurant_id = restaurant_id
    review.reviewer_email = reviewer_email
    review.score = score
    review.comments = comments
    review.price_subscore = price_subscore
    review.salsa_subscore = salsa_subscore
    review.tortillas_subscore = tortillas_subscore
    review.condiments_subscore = condiments_subscore
    review.ambiance_subscore = ambiance_subscore
    review.flavor_subscore = flavor_subscore
    review.service_subscore = service_subscore
    review.homemade_tortillas_yn = homemade_tortillas_yn
    review.other_notes = other_notes

    upload_dir = "/tmp/uploads"
    if photo_file_1:
        filename = f"{uuid.uuid4().hex}_{photo_file_1.filename}"
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, filename)
        with open(path, "wb") as buf:
            shutil.copyfileobj(photo_file_1.file, buf)
        if review.photo1:
            db.delete(review.photo1)
        review.photo1 = Photo(url=f"/static/uploads/{filename}")
        db.add(review.photo1)
    if photo_file_2:
        filename = f"{uuid.uuid4().hex}_{photo_file_2.filename}"
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, filename)
        with open(path, "wb") as buf:
            shutil.copyfileobj(photo_file_2.file, buf)
        if review.photo2:
            db.delete(review.photo2)
        review.photo2 = Photo(url=f"/static/uploads/{filename}")
        db.add(review.photo2)
    db.commit()
    return RedirectResponse(url="/admin/reviews", status_code=status.HTTP_303_SEE_OTHER)

# LIST – show all restaurants
@router.get("/restaurants", response_class=HTMLResponse)
async def list_restaurants(request: Request, db: Session = Depends(get_db)):
    restaurants = db.query(Restaurant).all()
    return HTMLResponse(
        content=jinja_templates.get_template("admin_restaurants.html").render(
            request=request, restaurants=restaurants
        )
    )
# NEW – display empty form
@router.get("/restaurants/new", response_class=HTMLResponse)
async def new_restaurant_form(request: Request):
    return HTMLResponse(
        content=jinja_templates.get_template("admin_restaurant_form.html").render(
            request=request,
            action="/admin/restaurants",
            method="POST",
            restaurant=None,
        )
    )
# CREATE – process form submission
@router.post("/restaurants")
async def create_restaurant(
    name: str = Form(...),
    address: str = Form(...),
    google_place_url: str = Form(None),
    google_map_url: str = Form(None),
    website_url: str = Form(None),
    yelp_url: str = Form(None),
    visited_yn: bool = Form(False),
    avg_score: float = Form(None),          # NEW
    review_count: int = Form(None),         # NEW
    db: Session = Depends(get_db),
):
    new_rest = Restaurant(
        name=name,
        address=address,
        google_place_url=google_place_url,
        google_map_url=google_map_url,
        website_url=website_url,
        yelp_url=yelp_url,
        visited_yn=visited_yn,
        avg_score=avg_score,               # NEW
        review_count=review_count,         # NEW
    )
    db.add(new_rest)
    db.commit()
    db.refresh(new_rest)
    return RedirectResponse(url="/admin/restaurants", status_code=status.HTTP_303_SEE_OTHER)

# EDIT – display pre‑filled form
@router.get("/restaurants/{restaurant_id}/edit", response_class=HTMLResponse)
async def edit_restaurant_form(restaurant_id: int, request: Request, db: Session = Depends(get_db)):
    restaurant = db.query(Restaurant).filter(Restaurant.id == restaurant_id).first()
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    return HTMLResponse(
        content=jinja_templates.get_template("admin_restaurant_form.html").render(
            request=request,
            action=f"/admin/restaurants/{restaurant_id}",
            method="POST",
            restaurant=restaurant,
        )
    )

# ---------- UPDATE ----------
@router.post("/restaurants/{restaurant_id}")
async def update_restaurant(
    restaurant_id: int,
    name: str = Form(...),
    address: str = Form(...),
    google_place_url: str = Form(None),
    google_map_url: str = Form(None),
    website_url: str = Form(None),
    yelp_url: str = Form(None),
    visited_yn: bool = Form(False),
    avg_score: float = Form(None),
    review_count: int = Form(None),
    db: Session = Depends(get_db),
):
    restaurant = db.query(Restaurant).filter(Restaurant.id == restaurant_id).first()
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    # Update fields …
    restaurant.name = name
    restaurant.address = address
    restaurant.google_place_url = google_place_url
    restaurant.google_map_url = google_map_url
    restaurant.website_url = website_url
    restaurant.yelp_url = yelp_url
    restaurant.visited_yn = visited_yn
    restaurant.avg_score = avg_score
    restaurant.review_count = review_count
    db.commit()
    # **Redirect back to the restaurant list**
    return RedirectResponse(
        url="/admin/restaurants",
        status_code=status.HTTP_303_SEE_OTHER
    )
# ---------- DELETE ----------
@router.post("/restaurants/{restaurant_id}/delete")
async def delete_restaurant(restaurant_id: int, db: Session = Depends(get_db)):
    restaurant = db.query(Restaurant).filter(Restaurant.id == restaurant_id).first()
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    db.delete(restaurant)
    db.commit()
    return RedirectResponse(url="/admin/restaurants", status_code=status.HTTP_303_SEE_OTHER)

# --------------------------------------------------------------
#  Generate the MyMaps Excel file (with hyperlinks)
# --------------------------------------------------------------
@router.get("/generate_xlsx")
async def generate_xlsx(db: Session = Depends(get_db)):
    """
    Build Dalton_Tacos.xlsx from the `restaurants` table.
    • URL columns become clickable hyperlinks.
    • The `review_url` column shows “Rate This Place” linking to the
      public review form for that restaurant.
    """
    # 1️⃣ Pull all restaurant records
    restaurants = db.query(Restaurant).all()
    # 2️⃣ Create workbook and header row
    wb = Workbook()
    ws = wb.active
    ws.title = "Dalton Tacos"
    headers = [
        "id",
        "name",
        "address",
        "google_place_url",
        "google_map_url",
        "website_url",
        "yelp_url",
        "visited_yn",
        "avg_score",
        "review_count",
        "review_url",
    ]
    ws.append(headers)
    # Bold the header row for readability
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    # ------------------------------------------------------
    # 3️⃣ Fill rows and add hyperlinks
    # ------------------------------------------------------
    for r in restaurants:
        # Append raw data (empty strings for None values)
        ws.append([
            r.id,
            r.name or "",
            r.address or "",
            r.google_place_url or "",
            r.google_map_url or "",
            r.website_url or "",
            r.yelp_url or "",
            r.visited_yn,
            r.avg_score,
            r.review_count,
            "",   # placeholder for review URL – will be set below
        ])
        cur_row = ws.max_row
        # ---- 3️⃣a Hyperlink the URL columns ----
        # Column indices are 1‑based; map each to its URL value.
        url_map = {
            4: r.google_place_url,
            5: r.google_map_url,
            6: r.website_url,
            7: r.yelp_url,
        }
        for col_idx, url in url_map.items():
            if url:  # only create a link if the URL isn't empty
                cell = ws.cell(row=cur_row, column=col_idx)
                cell.hyperlink = url
                cell.style = "Hyperlink"
        # ---- 3️⃣b Add hyperlink for the “Rate This Place” cell ----
        review_cell = ws.cell(row=cur_row, column=11)   # 11th column = review_url
        review_form_url = f"https://restaurant-system-gb7u.onrender.com/public/review_form?restaurant_id={r.id}"
        review_cell.value = review_form_url
        review_cell.hyperlink = review_form_url
        review_cell.style = "Hyperlink"
    # ------------------------------------------------------
    # 4️⃣ Stream the workbook back to the client
    # ------------------------------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Dalton_Tacos.xlsx"},
    )

